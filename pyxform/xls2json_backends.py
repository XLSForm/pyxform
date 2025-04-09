"""
XLS-to-dict and csv-to-dict are essentially backends for xls2json.
"""

import csv
import datetime
import re
from collections.abc import Callable, Iterable, Sequence
from dataclasses import dataclass
from enum import Enum
from io import BytesIO, IOBase, StringIO
from os import PathLike
from pathlib import Path
from typing import Any, BinaryIO
from zipfile import BadZipFile

from openpyxl import open as pyxl_open
from openpyxl.cell import Cell as pyxlCell
from openpyxl.reader.excel import ExcelReader
from openpyxl.workbook import Workbook as pyxlWorkbook
from openpyxl.worksheet.worksheet import Worksheet as pyxlWorksheet
from xlrd import XL_CELL_BOOLEAN, XL_CELL_DATE, XL_CELL_NUMBER, XLRDError
from xlrd import open_workbook as xlrd_open
from xlrd.book import Book as xlrdBook
from xlrd.sheet import Cell as xlrdCell
from xlrd.sheet import Sheet as xlrdSheet
from xlrd.xldate import XLDateAmbiguous, xldate_as_tuple

from pyxform import constants
from pyxform.errors import PyXFormError, PyXFormReadError

aCell = xlrdCell | pyxlCell
XL_DATE_AMBIGOUS_MSG = (
    "The xls file provided has an invalid date on the %s sheet, under"
    " the %s column on row number %s"
)
RE_WHITESPACE = re.compile(r"( )+")


@dataclass(slots=True)
class DefinitionData:
    # XLSForm definition sheets.
    # survey is optional to allow processing to proceed to warnings / spell checks.
    survey: Sequence[dict[str, str]] | None = None
    survey_header: Sequence[dict[str, Any]] | None = None
    choices: Sequence[dict[str, str]] | None = None
    choices_header: Sequence[dict[str, Any]] | None = None
    settings: Sequence[dict[str, str]] | None = None
    settings_header: Sequence[dict[str, Any]] | None = None
    external_choices: Sequence[dict[str, str]] | None = None
    external_choices_header: Sequence[dict[str, Any]] | None = None
    entities: Sequence[dict[str, str]] | None = None
    entities_header: Sequence[dict[str, Any]] | None = None
    osm: Sequence[dict[str, str]] | None = None
    osm_header: Sequence[dict[str, Any]] | None = None

    # Extra metadata.
    sheet_names: Sequence[str] | None = None
    fallback_form_name: str | None = None


def _list_to_dict_list(list_items):
    """
    Takes a list and creates a dict with the list values as keys.
    Returns a list of the created dict or an empty list
    """
    if list_items:
        return [{str(i): None for i in list_items}]
    return []


def trim_trailing_empty(a_list: list, n_empty: int) -> list:
    """
    Trim trailing empty columns or rows. Avoids `[:-0] == []`, and unnecessary list copy.
    """
    if 0 < n_empty:
        offset = len(a_list) - n_empty
        a_list = a_list[:offset]
    return a_list


def get_excel_column_headers(first_row: Iterable[str | None]) -> list[str | None]:
    """Get column headers from the first row; stop if there's a run of empty columns."""
    max_adjacent_empty_columns = 20
    column_header_list = []
    adjacent_empty_cols = 0
    for column_header in first_row:
        if is_empty(column_header):
            # Preserve column order (will filter later)
            column_header_list.append(None)
            # After a run of empty cols, assume we've reached the end of the data.
            if max_adjacent_empty_columns == adjacent_empty_cols:
                break
            adjacent_empty_cols += 1
        else:
            adjacent_empty_cols = 0
            # Check for duplicate column headers.
            if column_header in column_header_list:
                raise PyXFormError(f"Duplicate column header: {column_header}")
            # Strip whitespaces from the header.
            clean_header = RE_WHITESPACE.sub(" ", column_header.strip())
            column_header_list.append(clean_header)

    return trim_trailing_empty(column_header_list, adjacent_empty_cols)


def get_excel_rows(
    headers: Iterable[str | None],
    rows: Iterable[tuple[aCell, ...]],
    cell_func: Callable[[aCell, int, str], Any],
) -> list[dict[str, Any]]:
    """Get rows of cleaned data; stop if there's a run of empty rows."""
    max_adjacent_empty_rows = 60
    col_header_enum = list(enumerate(headers))
    adjacent_empty_rows = 0
    result_rows = []
    for row_n, row in enumerate(rows):
        row_dict = {}
        for col_n, key in col_header_enum:
            if key is None:
                continue
            try:
                cell = row[col_n]
                if not is_empty(cell.value):
                    row_dict[key] = cell_func(cell, row_n, key)
            except IndexError:
                pass  # rows may not have values for every column

        if 0 == len(row_dict):
            # After a run of empty rows, assume we've reached the end of the data.
            if max_adjacent_empty_rows == adjacent_empty_rows:
                break
            adjacent_empty_rows += 1
        else:
            adjacent_empty_rows = 0

        # There may be some empty rows amongst the XLSForm data. These are included
        # so that any warning messages that mention row numbers are accurate.
        result_rows.append(row_dict)

    return trim_trailing_empty(result_rows, adjacent_empty_rows)


def xls_to_dict(path_or_file):
    """
    Return a Python dictionary with a key for each worksheet
    name. For each sheet there is a list of dictionaries, each
    dictionary corresponds to a single row in the worksheet. A
    dictionary has keys taken from the column headers and values
    equal to the cell value for that row and column.
    All the keys and leaf elements are unicode text.
    """

    def xls_clean_cell(
        wb: xlrdBook, wb_sheet: xlrdSheet, cell: xlrdCell, row_n: int, col_key: str
    ) -> str | None:
        value = cell.value
        if isinstance(value, str):
            value = value.strip()
        if not is_empty(value):
            try:
                return xls_value_to_unicode(value, cell.ctype, wb.datemode)
            except XLDateAmbiguous as date_err:
                raise PyXFormError(
                    XL_DATE_AMBIGOUS_MSG % (wb_sheet.name, col_key, row_n)
                ) from date_err

        return None

    def xls_to_dict_normal_sheet(wb: xlrdBook, wb_sheet: xlrdSheet):
        # XLS format: max cols 256, max rows 65536
        first_row = (c.value for c in next(wb_sheet.get_rows(), []))
        headers = get_excel_column_headers(first_row=first_row)
        row_iter = (
            tuple(wb_sheet.cell(r, c) for c in range(len(headers)))
            for r in range(1, wb_sheet.nrows)
        )

        # Inject wb/sheet as closure since functools.partial isn't typing friendly.
        def clean_func(cell: xlrdCell, row_n: int, col_key: str) -> str | None:
            return xls_clean_cell(
                wb=wb, wb_sheet=wb_sheet, cell=cell, row_n=row_n, col_key=col_key
            )

        rows = get_excel_rows(headers=headers, rows=row_iter, cell_func=clean_func)
        column_header_list = [key for key in headers if key is not None]
        return rows, _list_to_dict_list(column_header_list)

    def process_workbook(wb: xlrdBook):
        result_book = {"sheet_names": []}
        for wb_sheet in wb.sheets():
            # Note original in sheet_names for spelling check.
            result_book["sheet_names"].append(wb_sheet.name)
            sheet_name = wb_sheet.name.lower()
            # Do not process sheets that have nothing to do with XLSForm.
            if sheet_name not in constants.SUPPORTED_SHEET_NAMES:
                if len(wb.sheets()) == 1:
                    (
                        result_book[constants.SURVEY],
                        result_book[f"{constants.SURVEY}_header"],
                    ) = xls_to_dict_normal_sheet(wb=wb, wb_sheet=wb_sheet)
                else:
                    continue
            else:
                (
                    result_book[sheet_name],
                    result_book[f"{sheet_name}_header"],
                ) = xls_to_dict_normal_sheet(wb=wb, wb_sheet=wb_sheet)
        return result_book

    try:
        wb_file = get_definition_data(definition=path_or_file)
        workbook = xlrd_open(file_contents=wb_file.data.getvalue())
        try:
            return process_workbook(wb=workbook)
        finally:
            workbook.release_resources()
    except (AttributeError, TypeError, XLRDError) as read_err:
        raise PyXFormReadError(f"Error reading .xls file: {read_err}") from read_err


def xls_value_to_unicode(value, value_type, datemode) -> str:
    """
    Take a xls formatted value and try to make a unicode string representation.
    """
    if value_type == XL_CELL_BOOLEAN:
        return "TRUE" if value else "FALSE"
    elif value_type == XL_CELL_NUMBER:
        # Try to display as an int if possible.
        int_value = int(value)
        if int_value == value:
            return str(int_value)
        else:
            return str(value)
    elif value_type is XL_CELL_DATE:
        # Warn that it is better to single quote as a string.
        # error_location = cellFormatString % (ss_row_idx, ss_col_idx)
        # raise Exception(
        #   "Cannot handle excel formatted date at " + error_location)
        datetime_or_time_only = xldate_as_tuple(value, datemode)
        if datetime_or_time_only[:3] == (0, 0, 0):
            # must be time only
            return str(datetime.time(*datetime_or_time_only[3:]))
        return str(datetime.datetime(*datetime_or_time_only))
    else:
        # ensure unicode and replace nbsp spaces with normal ones
        # to avoid this issue:
        # https://github.com/modilabs/pyxform/issues/83
        return str(value).replace(chr(160), " ")


def xlsx_to_dict(path_or_file):
    """
    Return a Python dictionary with a key for each worksheet
    name. For each sheet there is a list of dictionaries, each
    dictionary corresponds to a single row in the worksheet. A
    dictionary has keys taken from the column headers and values
    equal to the cell value for that row and column.
    All the keys and leaf elements are strings.
    """

    def xlsx_clean_cell(cell: pyxlCell, row_n: int, col_key: str) -> str | None:
        value = cell.value
        if isinstance(value, str):
            value = value.strip()
        if not is_empty(value):
            return xlsx_value_to_str(value)

        return None

    def xlsx_to_dict_normal_sheet(sheet: pyxlWorksheet):
        # XLSX format: max cols 16384, max rows 1048576
        first_row = (c.value for c in next(sheet.rows, []))
        headers = get_excel_column_headers(first_row=first_row)
        row_iter = sheet.iter_rows(min_row=2, max_col=len(headers))
        rows = get_excel_rows(headers=headers, rows=row_iter, cell_func=xlsx_clean_cell)
        column_header_list = [key for key in headers if key is not None]
        return rows, _list_to_dict_list(column_header_list)

    def process_workbook(wb: pyxlWorkbook):
        result_book = {"sheet_names": []}
        for sheetname in wb.sheetnames:
            # Note original in sheet_names for spelling check.
            result_book["sheet_names"].append(sheetname)
            sheet_name = sheetname.lower()
            # Do not process sheets that have nothing to do with XLSForm.
            if sheet_name not in constants.SUPPORTED_SHEET_NAMES:
                if len(wb.sheetnames) == 1:
                    (
                        result_book[constants.SURVEY],
                        result_book[f"{constants.SURVEY}_header"],
                    ) = xlsx_to_dict_normal_sheet(wb[sheetname])
                else:
                    continue
            else:
                (
                    result_book[sheet_name],
                    result_book[f"{sheet_name}_header"],
                ) = xlsx_to_dict_normal_sheet(wb[sheetname])
        return result_book

    try:
        wb_file = get_definition_data(definition=path_or_file)
        reader = ExcelReader(wb_file.data, read_only=True, data_only=True)
        reader.read()
        try:
            return process_workbook(wb=reader.wb)
        finally:
            reader.wb.close()
            reader.archive.close()
    except (BadZipFile, KeyError, OSError, TypeError) as read_err:
        raise PyXFormReadError(f"Error reading .xlsx file: {read_err}") from read_err


def xlsx_value_to_str(value) -> str:
    """
    Take a xls formatted value and try to make a string representation.
    """
    if value is True:
        return "TRUE"
    elif value is False:
        return "FALSE"
    elif isinstance(value, float) and value.is_integer():
        # Try to display as an int if possible.
        return str(int(value))
    elif isinstance(value, int | datetime.datetime | datetime.time):
        return str(value)
    else:
        # ensure unicode and replace nbsp spaces with normal ones
        # to avoid this issue:
        # https://github.com/modilabs/pyxform/issues/83
        value = str(value)
        if chr(160) in value:
            return value.replace(chr(160), " ")
        else:
            return value


def is_empty(value):
    if value is None:
        return True
    elif isinstance(value, str):
        if not value or value.isspace():
            return True

    return False


def get_cascading_json(sheet_list, prefix, level):
    return_list = []
    for row in sheet_list:
        if "stopper" in row:
            if row["stopper"] == level:
                # last element's name IS the prefix; doesn't need level
                return_list[-1]["name"] = prefix
                return return_list
            else:
                continue
        elif "lambda" in row:

            def replace_prefix(d, prefix):
                for k, v in d.items():
                    if isinstance(v, str):
                        d[k] = v.replace("$PREFIX$", prefix)
                    elif isinstance(v, dict):
                        d[k] = replace_prefix(v, prefix)
                    elif isinstance(v, list):
                        d[k] = (replace_prefix(x, prefix) for x in v)
                return d

            return_list.append(replace_prefix(row["lambda"], prefix))
    raise PyXFormError(
        "Found a cascading_select "
        + level
        + ", but could not find "
        + level
        + "in cascades sheet."
    )


def csv_to_dict(path_or_file):
    def first_column_as_sheet_name(row):
        if len(row) == 0:
            return None, None
        elif len(row) == 1:
            return row[0], None
        else:
            sheet_name = row[0].strip()
            content = [str(v).strip() for v in row[1:]]
            if sheet_name == "":
                sheet_name = None
            if not any(c != "" for c in content):
                # content is a list of empty strings
                content = None
            return sheet_name, content

    def process_csv_data(rd):
        _dict = {"sheet_names": []}
        sheet_name = None
        current_headers = None
        for row in rd:
            maybe_sheet_name, content = first_column_as_sheet_name(row)
            if maybe_sheet_name is not None:
                sheet_name = maybe_sheet_name
                if sheet_name and sheet_name not in _dict:
                    _dict["sheet_names"].append(sheet_name)
                    sheet_name = sheet_name.lower()
                    _dict[str(sheet_name)] = []
                current_headers = None
            if content is not None:
                if current_headers is None:
                    current_headers = content
                    _dict[f"{sheet_name}_header"] = _list_to_dict_list(current_headers)
                else:
                    _d = {
                        k: v
                        for k, v in zip(current_headers, content, strict=False)
                        if v != ""
                    }
                    _dict[sheet_name].append(_d)
        return _dict

    try:
        csv_data = get_definition_data(definition=path_or_file)
        csv_str = csv_data.data.getvalue().decode("utf-8")
        if not is_csv(data=csv_str):
            raise PyXFormError("The input data does not appear to be a valid XLSForm.")  # noqa: TRY301
        reader = csv.reader(StringIO(initial_value=csv_str, newline=""))
        return process_csv_data(rd=reader)
    except (AttributeError, PyXFormError) as read_err:
        raise PyXFormReadError(f"Error reading .csv file: {read_err}") from read_err


"""
I want the ability to go:

    xls => csv
so that we can go:
    xls => csv => survey

and some day:
    csv => xls

"""


def convert_file_to_csv_string(path):
    """
    This will open a csv or xls file and return a CSV in the format:
        sheet_name1
        ,col1,col2
        ,r1c1,r1c2
        ,r2c1,r2c2
        sheet_name2
        ,col1,col2
        ,r1c1,r1c2
        ,r2c1,r2c2

    Currently, it processes csv files and xls files to ensure consistent
    csv delimiters, etc. for tests.
    """
    if path.endswith(".csv"):
        imported_sheets = csv_to_dict(path)
    else:
        imported_sheets = xls_to_dict(path)
    foo = StringIO(newline="")
    writer = csv.writer(foo, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL)
    for sheet_name, rows in imported_sheets.items():
        if sheet_name == "sheet_names":
            continue
        writer.writerow([sheet_name])
        out_keys = []
        out_rows = []
        for row in rows:
            out_row = []
            for key in row.keys():
                if key not in out_keys:
                    out_keys.append(key)
            for out_key in out_keys:
                out_row.append(row.get(out_key, None))
            out_rows.append(out_row)
        writer.writerow([None, *out_keys])
        for out_row in out_rows:
            writer.writerow([None, *out_row])
    return foo.getvalue()


def sheet_to_csv(workbook_path, csv_path, sheet_name):
    if workbook_path.endswith(".xls"):
        return xls_sheet_to_csv(workbook_path, csv_path, sheet_name)
    else:
        return xlsx_sheet_to_csv(workbook_path, csv_path, sheet_name)


def xls_sheet_to_csv(workbook_path, csv_path, sheet_name):
    wb = xlrd_open(workbook_path)
    try:
        sheet = wb.sheet_by_name(sheet_name)
    except XLRDError:
        return False
    if not sheet or sheet.nrows < 2:
        return False
    with open(csv_path, mode="w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_ALL)
        mask = [v and len(v.strip()) > 0 for v in sheet.row_values(0)]
        for row_idx in range(sheet.nrows):
            csv_data = []
            try:
                for v, m in zip(sheet.row(row_idx), mask, strict=False):
                    if m:
                        value = v.value
                        value_type = v.ctype
                        data = xls_value_to_unicode(value, value_type, wb.datemode)
                        # clean the values of leading and trailing whitespaces
                        data = data.strip()
                        csv_data.append(data)
            except TypeError:
                continue
            writer.writerow(csv_data)

    return True


def xlsx_sheet_to_csv(workbook_path, csv_path, sheet_name):
    wb = pyxl_open(workbook_path, read_only=True, data_only=True)
    try:
        sheet = wb[sheet_name]
    except KeyError:
        return False

    with open(csv_path, mode="w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_ALL)
        mask = [not is_empty(cell.value) for cell in sheet[1]]
        for row in sheet.rows:
            csv_data = []
            try:
                for v, m in zip(row, mask, strict=False):
                    if m:
                        data = xlsx_value_to_str(v.value)
                        # clean the values of leading and trailing whitespaces
                        data = data.strip()
                        csv_data.append(data)
            except TypeError:
                continue
            writer.writerow(csv_data)
    wb.close()
    return True


MD_COMMENT = re.compile(r"^\s*#")
MD_COMMENT_INLINE = re.compile(r"^(.*)(#[^|]+)$")
MD_CELL = re.compile(r"\s*\|(.*)\|\s*")
MD_SEPARATOR = re.compile(r"^[\|-]+$")
MD_PIPE_OR_ESCAPE = re.compile(r"(?<!\\)\|")


def _md_strp_cell(cell: str) -> str:
    if not cell or cell.isspace():
        return None
    return cell.strip().replace(r"\|", "|")


def _md_table_to_ss_structure(mdstr: str) -> dict[str, list[tuple[str, ...]]]:
    sheet_name = False
    sheet_arr = False
    sheets = {}
    for line in mdstr.split("\n"):
        if re.match(MD_COMMENT, line):
            # ignore lines which start with pound sign
            continue
        elif re.match(MD_COMMENT_INLINE, line):
            # keep everything before the # outside of the last occurrence of |
            line = re.match(MD_COMMENT_INLINE, line).groups()[0]
        match = re.match(MD_CELL, line)
        if match:
            mtchstr = match.groups()[0]
            if not re.match(MD_SEPARATOR, mtchstr):
                row_split = re.split(MD_PIPE_OR_ESCAPE, mtchstr)
                first_col = _md_strp_cell(row_split[0])
                row = tuple(_md_strp_cell(c) for c in row_split[1:])
                if first_col is None and row is None:
                    continue
                if first_col is not None:
                    if sheet_arr:
                        sheets[sheet_name] = sheet_arr
                    sheet_arr = []
                    sheet_name = first_col
                if sheet_name and any(c is not None for c in row):
                    sheet_arr.append(row)
            sheets[sheet_name] = sheet_arr

    return sheets


def md_to_dict(md: str | BytesIO):
    def list_to_dicts(arr):
        return [
            {arr[0][i]: v for i, v in enumerate(row) if v not in {None, ""}}
            for row in arr[1:]
        ]

    def process_md_data(md_: str):
        result_book = {"sheet_names": []}
        ss_structure = _md_table_to_ss_structure(md_)
        for sheet, contents in ss_structure.items():
            # Note original in sheet_names for spelling check.
            result_book["sheet_names"].append(sheet)
            sheet_name = sheet.lower()
            # Do not process sheets that have nothing to do with XLSForm.
            if sheet_name not in constants.SUPPORTED_SHEET_NAMES:
                if len(ss_structure) == 1:
                    result_book[constants.SURVEY] = list_to_dicts(contents)
                    result_book[f"{constants.SURVEY}_header"] = _list_to_dict_list(
                        contents[0]
                    )
                else:
                    continue
            else:
                result_book[sheet_name] = list_to_dicts(contents)
                result_book[f"{sheet_name}_header"] = _list_to_dict_list(contents[0])
        return result_book

    try:
        md_data = get_definition_data(definition=md)
        md_str = md_data.data.getvalue().decode("utf-8")
        if not is_markdown_table(data=md_str):
            raise PyXFormError("The input data does not appear to be a valid XLSForm.")  # noqa: TRY301
        return process_md_data(md_=md_str)
    except (AttributeError, PyXFormError, TypeError) as read_err:
        raise PyXFormReadError(f"Error reading .md file: {read_err}") from read_err


def md_table_to_workbook(mdstr: str) -> pyxlWorkbook:
    """
    Convert Markdown table string to an openpyxl.Workbook. Call wb.save() to persist.
    """
    md_data = _md_table_to_ss_structure(mdstr=mdstr)
    wb = pyxlWorkbook(write_only=True)
    for key, rows in md_data.items():
        sheet = wb.create_sheet(title=key)
        for r in rows:
            sheet.append(r)
    return wb


def count_characters_limit(data: str, find: str, limit: int) -> int:
    count = 0
    for c in data:
        if c == find:
            count += 1
            if count == limit:
                break
    return count


def is_markdown_table(data: str) -> bool:
    """
    Does the string look like a markdown table? Checks the first 5KB.

    A minimal form with one question requires 10 pipe characters:

    | survey |
    |        | type    | name |
    |        | integer | a    |

    Minimum to parse at all is 5 pipes.

    | survey |
    |        | foo |

    :param data: The data to check.
    """
    return 5 <= count_characters_limit(data[:5000], "|", 5)


def is_csv(data: str) -> bool:
    """
    Does the string look like a CSV? Checks the first 5KB.

    A minimal form with one question requires 4 comma characters:

    "survey"
    ,"type","name"
    ,"integer","a"

    :param data: The data to check.
    """
    return 4 <= count_characters_limit(data[:5000], ",", 4)


class SupportedFileTypes(Enum):
    xlsx = ".xlsx"
    xlsm = ".xlsm"
    xls = ".xls"
    md = ".md"
    csv = ".csv"

    @staticmethod
    def get_processors():
        return {
            SupportedFileTypes.xlsx: xlsx_to_dict,
            SupportedFileTypes.xlsm: xlsx_to_dict,
            SupportedFileTypes.xls: xls_to_dict,
            SupportedFileTypes.md: md_to_dict,
            SupportedFileTypes.csv: csv_to_dict,
        }


@dataclass(slots=True)
class Definition:
    data: BytesIO
    file_type: SupportedFileTypes | None
    file_path_stem: str | None


def definition_to_dict(
    definition: str | PathLike[str] | bytes | BytesIO | IOBase | Definition,
    file_type: str | None = None,
) -> DefinitionData:
    """
    Convert raw definition data to a dict ready for conversion to a XForm.

    :param definition: XLSForm definition data.
    :param file_type: If provided, attempt parsing the data only as this type. Otherwise,
      parsing of supported data types will be attempted until one of them succeeds.
    :return:
    """
    supported = f"Must be one of: {', '.join(t.value for t in SupportedFileTypes)}"
    processors = SupportedFileTypes.get_processors()
    if file_type is not None:
        try:
            ft = SupportedFileTypes(file_type)
        except ValueError as err:
            raise PyXFormError(
                f"Argument 'file_type' is not a supported type. {supported}"
            ) from err
        else:
            processors = {ft: processors[ft]}

    for func in processors.values():
        try:
            return DefinitionData(
                fallback_form_name=definition.file_path_stem, **func(definition)
            )
        except PyXFormReadError:  # noqa: PERF203
            continue

    raise PyXFormError(
        f"Argument 'definition' was not recognized as a supported type. {supported}"
    )


def get_definition_data(
    definition: str | PathLike[str] | bytes | BytesIO | IOBase | Definition,
) -> Definition:
    """
    Get the form definition data from a path or bytes.

    :param definition: The path to the file to upload (string or PathLike), or the
        form definition in memory (string or bytes).
    """
    if isinstance(definition, Definition):
        return definition
    definition_data = None
    file_type = None
    file_path_stem = None

    # Read in data from paths, or failing that try to process the string.
    if isinstance(definition, str | PathLike):
        file_read = False
        try:
            file_path = Path(definition)

        except TypeError as err:
            raise PyXFormError(
                "Parameter 'definition' does not appear to be a valid file path."
            ) from err
        try:
            file_exists = file_path.is_file()
        except (FileNotFoundError, OSError):
            pass
        else:
            if file_exists:
                file_path_stem = file_path.stem
                try:
                    file_type = SupportedFileTypes(file_path.suffix)
                except ValueError:
                    # The suffix was not a useful hint but we can try to parse anyway.
                    pass
                definition = BytesIO(file_path.read_bytes())
                file_read = True
        if not file_read and isinstance(definition, str):
            definition = definition.encode("utf-8")

    # io.IOBase seems about at close as possible to the hint typing.BinaryIO.
    if isinstance(definition, bytes | BytesIO | IOBase):
        # Normalise to BytesIO.
        if isinstance(definition, bytes):
            definition_data = BytesIO(definition)
        elif isinstance(definition, BytesIO):  # BytesIO is a subtype of IOBase
            definition_data = definition
        else:
            definition_data = BytesIO(definition.read())

    return Definition(
        data=definition_data,
        file_type=file_type,
        file_path_stem=file_path_stem,
    )


def get_xlsform(
    xlsform: str | PathLike[str] | bytes | BytesIO | BinaryIO | dict,
    file_type: str | None = None,
) -> DefinitionData:
    if isinstance(xlsform, dict):
        workbook_dict = DefinitionData(**xlsform)
    else:
        definition = get_definition_data(definition=xlsform)
        if file_type is None:
            file_type = definition.file_type
        workbook_dict = definition_to_dict(definition=definition, file_type=file_type)
    return workbook_dict
